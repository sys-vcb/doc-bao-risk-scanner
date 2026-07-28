import os
from datetime import datetime
from typing import List, Optional
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import REPORTS_DIR
from app.models import RiskItem

def generate_excel_report(
    risk_items: List[RiskItem], 
    date_from: Optional[str] = None, 
    date_to: Optional[str] = None, 
    province: Optional[str] = None
) -> str:
    """
    Xuất báo cáo rủi ro doanh nghiệp ra file Excel (.xlsx) chuẩn 5 cột có định dạng sang trọng.
    """
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    # Đặt tên file
    prov_tag = f"_{province}" if province and province != "Tất cả" else ""
    date_tag = f"_{date_from}_den_{date_to}" if date_from and date_to else f"_{today_str}"
    filename = f"Bao_Cao_Rui_Ro{prov_tag}{date_tag}.xlsx"
    file_path = REPORTS_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo Cáo Rủi Ro"
    ws.views.sheetView[0].showGridLines = True

    # Palette Màu Slate Glassmorphism
    HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    TITLE_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    
    FONT_TITLE = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    FONT_SUBTITLE = Font(name="Calibri", size=10, italic=True, color="94A3B8")
    FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    FONT_DATA = Font(name="Calibri", size=10, color="0F172A")
    FONT_LINK = Font(name="Calibri", size=10, color="2563EB", underline="single")
    
    THIN_BORDER = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # 1. Dòng Tiêu đề Báo Cáo
    ws.merge_cells("A1:E1")
    cell_title = ws["A1"]
    cell_title.value = "BÁO CÁO CẢNH BÁO RỦI RO DOANH NGHIỆP / TỔ CHỨC / CÁ NHÂN"
    cell_title.font = FONT_TITLE
    cell_title.fill = TITLE_FILL
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Dòng Phụ đề Filter
    ws.merge_cells("A2:E2")
    cell_sub = ws["A2"]
    filter_desc = f"Khoảng thời gian: {date_from or 'Tất cả'} đến {date_to or 'Hiện tại'} | Khu vực: {province or 'Tất cả 6 tỉnh'} | Tổng số rủi ro: {len(risk_items)}"
    cell_sub.value = filter_desc
    cell_sub.font = FONT_SUBTITLE
    cell_sub.fill = TITLE_FILL
    cell_sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 10 # Dòng trống phân cách

    # 2. Tiêu đề 5 Cột chuẩn
    headers = [
        "1. Tên Doanh nghiệp / Cá nhân / Tổ chức",
        "2. Nội dung tóm tắt rủi ro (2-3 câu)",
        "3. Khu vực",
        "4. Loại rủi ro / Từ khóa",
        "5. Ngày tin tức & Link bài gốc"
    ]
    
    ws.row_dimensions[4].height = 28
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header_title
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # 3. Điền Dữ liệu các dòng
    start_row = 5
    if not risk_items:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
        empty_cell = ws.cell(row=start_row, column=1)
        empty_cell.value = "Không tìm thấy dữ liệu rủi ro nào thỏa mãn khoảng thời gian và khu vực đã chọn."
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        empty_cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
        ws.row_dimensions[start_row].height = 25
    else:
        for idx, item in enumerate(risk_items):
            current_row = start_row + idx
            ws.row_dimensions[current_row].height = 55 # Độ cao thích hợp cho tóm tắt 2-3 câu
            
            fill = ZEBRA_FILL if idx % 2 == 1 else PatternFill(fill_type=None)

            # Cột 1: Tên Doanh nghiệp / Cá nhân
            c1 = ws.cell(row=current_row, column=1, value=item.entity_name or "Chưa xác định")
            c1.alignment = Alignment(vertical="center", wrap_text=True)

            # Cột 2: Nội dung tóm tắt
            c2 = ws.cell(row=current_row, column=2, value=item.summary or "")
            c2.alignment = Alignment(vertical="center", wrap_text=True)

            # Cột 3: Khu vực
            c3 = ws.cell(row=current_row, column=3, value=item.province or "Toàn quốc")
            c3.alignment = Alignment(horizontal="center", vertical="center")

            # Cột 4: Loại rủi ro
            c4 = ws.cell(row=current_row, column=4, value=item.risk_type or "")
            c4.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Cột 5: Ngày & Hyperlink Bài gốc
            c5 = ws.cell(row=current_row, column=5)
            if item.source_url:
                c5.value = f"{item.published_date}\n(Xem bài gốc)"
                c5.hyperlink = item.source_url
                c5.font = FONT_LINK
            else:
                c5.value = item.published_date
                c5.font = FONT_DATA
            c5.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Áp dụng định dạng chung cho toàn bộ cells trong hàng
            for col_num in range(1, 6):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = THIN_BORDER
                if fill.fill_type and col_num != 5:
                    cell.fill = fill
                if col_num != 5:
                    cell.font = FONT_DATA

    # Cấu hình độ rộng tối ưu cho 5 cột
    col_widths = {
        1: 28,  # Tên Doanh nghiệp/Cá nhân
        2: 55,  # Nội dung tóm tắt
        3: 16,  # Khu vực
        4: 25,  # Loại rủi ro
        5: 22   # Ngày & Link
    }
    for col_num, width in col_widths.items():
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width

    wb.save(str(file_path))
    return str(file_path)
